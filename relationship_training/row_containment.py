import openpyxl as pxl
from sys import argv
from os import path, rename
from random import sample, random
import shutil
from math import floor
from hashlib import md5

EDGE_PATH = './row_containment_edges.csv'

def create_file():
    header = "A_file, A_sheet, B_file, B_sheet"
    with open(EDGE_PATH, "w") as f:
        f.write(header)
        f.write('\n')

""" Given an excel filepath 
create a row containment candidate, add the relationship to master file.
"""
def create_candidate(filepath):
    # copy file
    newpath = filepath[:-5] + 'CP' + filepath[-5:]
    shutil.copy(filepath, newpath)
    # modify copy for row containment.
    wb = pxl.load_workbook(newpath)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        delete_num = int(floor(random() * ws.max_row)) - 1
        deletions = sample(range(ws.max_row), delete_num)
        deletions = sorted(deletions, reverse=True)
        for d in deletions:
            ws.delete_rows(d,1)
    wb.save()
    # compute hash, rename file.
    dst = md5(newpath).hexdigest()[:16]
    dst = newpath.rsplit('/',1)[0] + dst + '.xlsx'
    rename(newpath, dst)
    # add edge
    A_file = filepath.rsplit('/',1)[1]
    B_file = dst.rsplit('/',1)[1]
    if not path.exists(EDGE_PATH):
        create_file()
    with open(EDGE_PATH, 'a+') as f:
        for name in wb.sheetnames:
            f.write(f'{A_file},{name},{B_file},{name}')
    
if __name__ == "__main__":
    path = argv[1]
    if path.exists(path):
        create_candidate(path)
    else:
        print(f"error: path {path} does not exist.")