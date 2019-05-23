
import xlrd
import openpyxl
import numpy as np
import cell
import keras
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import Feature_builder
from openpyxl.workbook import Workbook
from sklearn.neural_network import MLPClassifier
from tkinter import *
from tkinter import filedialog
import json
import pickle
import pandas
from keras.models import load_model



root = Tk()
fileName = filedialog.askopenfilenames(parent=root, title='Choose a file')
all_File = root.tk.splitlist(fileName)
dictionary = {}
final = []
for i in range(len(all_File)):
    file_url = all_File[i]
    file_list = file_url.split("/")
    final.append("/"+file_list[len(file_list)-2]+"/"+file_list[len(file_list)-1])
dictionary["url"] = final
print(all_File)
with open("fileName.json", 'w') as outfile:
    json.dump(dictionary, outfile)



feature_vector_list = []
position_list = []
for item in all_File:
    feature_vector_workbook = []
    position_in_workbook = []
    wb_obj = openpyxl.load_workbook(item,data_only=False)
    worksheet_list = wb_obj.get_sheet_names()
    for item in worksheet_list:
        sheet_obj = wb_obj[item]
        merge_Range = sheet_obj.merged_cells.ranges




        if len(merge_Range) != 0:
            for item in merge_Range:
                sheet_obj.unmerge_cells(str(item))

        feature_vector = []
        temp = []
        position = []

        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column
        for i in range(1, max_row+1):
            for j in range(1, max_col+1):
                temp = []
                cell_obj = sheet_obj.cell(row=i, column=j)
                c = cell.cell()
                print(cell_obj.value)
                print(type(cell_obj.value))
                c.setAttributes(cell_obj)

                if i != 1:
                    above_cell_obj = sheet_obj.cell(row=i-1, column=j)
                    c_a = cell.cell()
                    c_a.setAttributes(above_cell_obj)
                    c.setAboveNeighbor(c_a)

                below_cell_obj = sheet_obj.cell(row=i+1, column=j)
                c_b = cell.cell()
                c_b.setAttributes(below_cell_obj)
                c.setBelowNeighbour(c_b)
                if j != 1:
                    left_cell_obj = sheet_obj.cell(row=i, column=j-1)
                    c_f = cell.cell()
                    c_f.setAttributes(left_cell_obj)
                    c.setLeftNeighbour(c_f)

                right_cell_obj = sheet_obj.cell(row=i, column=j+1)
                c_r = cell.cell()
                c_r.setAttributes(right_cell_obj)
                c.setRightNeighbour(c_r)

                if c.get_Is_blank():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Bold_font():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Below_blank():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.has_merge_cell == False:
                    temp.append(0)
                else:
                    temp.append(1)

                if c.get_Above_alpha():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Left_align():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Right_blank():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Above_blank():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Above_num():
                    temp.append(1)
                else:
                    temp.append(0)


                if c.get_Above_alphanum():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Right_align():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Underline_font():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Below_num():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Left_alpha():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Above_in_header():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Left_num():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_All_small():
                    temp.append(1)
                else:
                    temp.append(0)

                if c .get_Is_alpha():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Right_num():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Text_in_header():
                    temp.append(1)
                else:
                    temp.append(0)

                if c.get_Is_num():
                    temp.append(1)
                else:
                    temp.append(0)

                feature_vector.append(temp)  # feature vector of one cell -- temp
                position.append((i, j))

        feature_vector = np.array(feature_vector)  # all feature vector in one worksheet
        position = np.array(position)
        if len(feature_vector) != 0:
            feature_vector_workbook.append(feature_vector)
        if len(position) != 0:
            position_in_workbook.append(position)

    feature_vector_list.append(feature_vector_workbook)
    position_list.append(position_in_workbook)


pinkFill = PatternFill(start_color='f9636c',
                       end_color='f9636c',
                       fill_type='solid')

yellowFill = PatternFill(start_color='f9ff00',
                         end_color='f9ff00',
                         fill_type='solid')

greenFill = PatternFill(start_color='39ff00',
                        end_color='39ff00',
                        fill_type='solid')


deepgreenFill = PatternFill(start_color='28e107',
                            end_color='28e107',
                            fill_type='solid')

blueFill = PatternFill(start_color='00ecff',
                       end_color='00ecff',
                       fill_type='solid')


workbook_dict = {}

model = load_model('model0.7.h5')
for i in range(len(feature_vector_list)):
    wb_obj = openpyxl.load_workbook(all_File[i])
    worksheet_list = wb_obj.get_sheet_names()
    item1 = feature_vector_list[i]
    item2 = position_list[i]
    worksheet_dict = {}
    for j in range(len(item1)):

        dict = {}
        sheet_obj = wb_obj[worksheet_list[j]]
        m = model.predict(item1[j])


        add = item2[j]
        for k in range(len(m)):
            label = np.argmax(m[k])
            if label not in dict.keys():
                dict[label] = []
                dict[label].append(add[k].tolist())
            else:
                dict[label].append(add[k].tolist())
        worksheet_key = "Worksheet"+str(j)
        worksheet_dict[worksheet_key] = dict
    workbook_key = "workbook" + str(i)
    workbook_dict[workbook_key] = worksheet_dict
with open("cell_label.json", 'w') as outfile:
    json.dump(workbook_dict, outfile)


        # for k in range(len(label)):
            if label == 4:
                sheet_obj.cell(row=add[k][0], column=add[k][1]).fill = pinkFill

            elif label == 0:
                sheet_obj.cell(row=add[k][0], column=add[k][1]).fill = yellowFill

            elif label == 3:
                sheet_obj.cell(row=add[k][0], column=add[k][1]).fill = blueFill

            elif label == 1:
                sheet_obj.cell(row=add[k][0], column=add[k][1]).fill = greenFill

            else:
                sheet_obj.cell(row=add[k][0], column=add[k][1]).fill = deepgreenFill

wb_obj.save(all_File[i])



#order of label [CH, DC, DE, DS, NDC]

