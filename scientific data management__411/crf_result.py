import pandas
import numpy as np
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border

colnames = ["is_blank", "bold_font","below_blank","has_merge_cell","above_alpha","left_align","right_blank","above_blank","above_num","above_alphanum","right_align","underline_font","below_num","left_alpha","above_in_header","left_num","all_small","is_alpha","right_num","text_in_header","is_num","label"]

data = pandas.read_csv("outputfile.txt", names=colnames,header=None, delimiter="\t")
data = pandas.DataFrame(data)
label = data["label"]
label = np.array(label)


wb_obj = openpyxl.load_workbook("TestFiles/00sumdat_copy.xlsx", data_only=False)
sheet_obj = wb_obj["Sheet1"]

pinkFill = PatternFill(start_color='f9636c',
                       end_color='f9636c',
                       fill_type='solid')

yellowFill = PatternFill(start_color='f9ff00',
                         end_color='f9ff00',
                         fill_type='solid')

greenFill = PatternFill(start_color='39ff00',
                        end_color='39ff00',
                        fill_type='solid')


deepgreenFill = PatternFill(start_color='28e107',
                            end_color='28e107',
                            fill_type='solid')

blueFill = PatternFill(start_color='00ecff',
                       end_color='00ecff',
                       fill_type='solid')


max_row = sheet_obj.max_row
max_col = sheet_obj.max_column

label = label.reshape(max_row,max_col)
for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        if label[i-1][j-1] == "NDC":
            sheet_obj.cell(row=i, column=j).fill = pinkFill

        elif label[i-1][j-1] == "CH":
            sheet_obj.cell(row=i, column=j).fill = yellowFill

        elif label[i-1][j-1] == "DS":
            sheet_obj.cell(row=i, column=j).fill = blueFill

        elif label[i-1][j-1] == "DC":
            sheet_obj.cell(row=i, column=j).fill = greenFill

        else:
            sheet_obj.cell(row=i, column=j).fill = deepgreenFill
wb_obj.save("TestFiles/00sumdat_copy.xlsx")