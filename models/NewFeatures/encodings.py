import glob
import openpyxl as pyxl
import pandas as pd
files = glob.glob('./training_spreadsheets/**/*.xlsx', recursive=True)
# print(files)
# exit()
train_df = pd.read_csv('./data_corpus/training_files/vanilla_training.csv', header=0)
worksheets = []
sheet_tokens = []
for f in files:
    try:
        px = pyxl.load_workbook(f)
        worksheets.append(px)
        # print(px.sheetnames)
        filename=f.split('/')[-1].replace(' ','_')
        tokens = [f"{filename}___{sheet.replace(' ','_')}___" for sheet in px.sheetnames]
        sheet_tokens += tokens
    except:
        continue
    
# print(train_df.file_name)
for token in sheet_tokens:
    # print(f'checking for {token}')
    if train_df['file_name'].str.contains(token).any():
        print(f'Y {token}')
    else:
        print(f'N {token}')
# print(sheet_tokens)
# print(sheets)