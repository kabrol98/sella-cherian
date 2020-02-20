import math
from collections import Counter
from os import path

import numpy as np
import openpyxl as xl

from components.cell_labeling.cell_compact import CellTagType, ContentType
from components.cell_labeling.cell_extended import CellExtended
from components.extract_column.extract_helper import ExtractHelper
from components.parse_files.metadata import ColumnMetaData


class CellLabeled:
    def __init__(self, tag, cell):
        self.tag = tag
        self.cell = cell


class Parser:
    def __init__(self, file_path: str, model):
        self.file_path = file_path
        self.model = model

    def file_validity_check(self) -> bool:
        if not path.exists(self.file_path):
            print("File does not exist")
            return False
        if not path.isfile(self.file_path):
            print("Path specified is not a file")
            return False
        if not (self.file_path.endswith(".xlsx") or self.file_path.endswith(".xls")):
            print("File must be either xlsx or xls")
            return False
        return True


    def set_rules(self, cells):
        result = []
        for col in cells:
            first_cell = None
            for d in col:
                if not d.cell.state["is_blank"]:
                    first_cell = d
                    break
            if first_cell is None:
                for d in col:
                    d.tag = CellTagType.NDC
                continue
            else:
                if first_cell.tag == CellTagType.CH:
                    pass
                else:
                    first_cell.tag = CellTagType.DS
            last_cell = None
            for d in reversed(col):
                if not d.cell.state["is_blank"]:
                    last_cell = d
                    break
            if last_cell is None:
                raise Exception("last cell must not be None after checking first cell")
            elif last_cell == first_cell:
                pass
            else:
                cell_tag = last_cell.tag
                if cell_tag == CellTagType.DE or cell_tag == CellTagType.CH:
                    pass
                else:
                    last_cell.tag = CellTagType.DE

        last_empty_row = -math.inf
        last_num_row = -math.inf
        cells_t = [list(x) for x in zip(*cells)]
        changed_rows = set()
        potential_string_rows = set()
        for i in range(len(cells_t)):
            row = cells_t[i]
            num_num_cells = 0
            tag_num_cells = []
            num_string_cells = 0
            for cell_labelled in row:
                if cell_labelled.cell.compact_cell.content_type == ContentType.NUMERIC:
                    num_num_cells += 1
                    if cell_labelled.tag != CellTagType.NDC:
                        tag_num_cells.append(cell_labelled.tag)
                elif cell_labelled.cell.compact_cell.content_type == ContentType.STRING:
                    num_string_cells += 1
            if num_num_cells > 0:
                most_common_tag = Counter(tag_num_cells).most_common(1)[0][0]
                for cell_labelled in row:
                    if cell_labelled.cell.compact_cell.content_type == ContentType.STRING:
                        cell_labelled.tag = most_common_tag
                        changed_rows.add(i)
                last_num_row = i
            elif num_num_cells == 0:
                if i == 0 or last_num_row == i - 1:
                    for cell_labelled in row:
                        if cell_labelled.cell.compact_cell.content_type == ContentType.STRING:
                            cell_labelled.tag = CellTagType.CH
                            changed_rows.add(i)
                    if last_num_row == i - 1:
                        # assume that numbers cannot be headers
                        for cell_labelled in cells_t[i-1]:
                            cell_labelled.tag = CellTagType.DE
                            changed_rows.add(i-1)
                else:
                    if i not in changed_rows:
                        potential_string_rows.add(i)

        for row_idx in potential_string_rows:
            isStringCol = True
            for iter_row_idx in range(row_idx + 1, len(cells_t)):
                if iter_row_idx - row_idx <= 3:
                    if iter_row_idx not in potential_string_rows:
                        isStringCol = False
                        break
                else:
                    break
            if isStringCol:
                row = cells_t[row_idx]
                for cell_labelled in row:
                    if cell_labelled.tag == CellTagType.CH or cell_labelled.tag == CellTagType.DS:
                        cell_labelled.tag = CellTagType.DC


        print('\n=======================================================================================\n' * 10)
        for col in cells:
            result.append(list(
                map(lambda temp: CellLabeled(tag=temp.tag, cell=temp.cell.compact_cell), col)
            ))
            for cell in col:
                print(cell.tag, cell.cell.compact_cell)
        return result


    def classify_cells(self, cells):
        result = []
        for col in cells:
            col_result = []
            for cell in col:
                col_result.append(cell.get_feature_vector() + 1)
            classification = self.model.predict([col_result], batch_size=1)
            tagged_result = []
            # print(classification)
            for idx, cell in enumerate(col):
                label = int(np.argmax(classification[idx][0][1:]))  # the first tag should not exist
                # print(label)
                # test_exit()
                if label < 0 or label > 5:
                    raise RuntimeError("Invalid Label")
                temp_dict = CellLabeled(tag=label, cell=cell)
                # if cell.compact_cell is not None:
                    # ['CH' 'DC' 'DE' 'DS' 'NDC']
                    # print(temp_dict.tag, temp_dict.cell.compact_cell, cell.get_feature_vector())
                # print(label)
                tagged_result.append(temp_dict)
            result.append(tagged_result)
        # test_exit()
        print("\n\n")
        return self.set_rules(result)

    def parse(self):
        result = self.file_validity_check()
        if not result:
            return None
        workbook = xl.load_workbook(self.file_path)
        sheet_names = workbook.get_sheet_names()
        columns = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            merge_range = worksheet.merged_cells.ranges
            if len(merge_range) != 0:
                for item in merge_range:
                    worksheet.unmerge_cells(str(item))
            max_row_num = worksheet.max_row
            max_col_num = worksheet.max_column
            virtual_worksheet = [x[:] for x in [[None] * max_row_num] * max_col_num]
            for j in range(1, max_col_num + 1):
                for i in range(1, max_row_num + 1):
                    virtual_i = i - 1
                    virtual_j = j - 1
                    cell = virtual_worksheet[virtual_j][virtual_i]
                    if cell is None:
                        cell = CellExtended(worksheet.cell(row=i, column=j))
                        virtual_worksheet[virtual_j][virtual_i] = cell
                    if virtual_i - 1 >= 0:
                        top_cell = virtual_worksheet[virtual_j][virtual_i - 1]
                        if top_cell is None:
                            top_cell = CellExtended(worksheet.cell(row=i - 1, column=j))
                            virtual_worksheet[virtual_j][virtual_i - 1] = top_cell
                        cell.apply_above_neighbor(top_cell)
                        top_cell.apply_below_neighbor(cell)
                    else:
                        cell.apply_above_neighbor(None)
                    if virtual_i + 1 < max_row_num:
                        bottom_cell = virtual_worksheet[virtual_j][virtual_i + 1]
                        if bottom_cell is None:
                            bottom_cell = CellExtended(worksheet.cell(row=i + 1, column=j))
                            virtual_worksheet[virtual_j][virtual_i + 1] = bottom_cell
                        cell.apply_below_neighbor(bottom_cell)
                        bottom_cell.apply_above_neighbor(cell)
                    else:
                        cell.apply_below_neighbor(None)
                    if virtual_j - 1 >= 0:
                        left_cell = virtual_worksheet[virtual_j - 1][virtual_i]
                        if left_cell is None:
                            left_cell = CellExtended(worksheet.cell(row=i, column=j - 1))
                            virtual_worksheet[virtual_j - 1][virtual_i] = left_cell
                        cell.apply_left_neighbor(left_cell)
                        left_cell.apply_right_neighbor(cell)
                    else:
                        cell.apply_left_neighbor(None)
                    if virtual_j + 1 < max_col_num:
                        right_cell = virtual_worksheet[virtual_j + 1][virtual_i]
                        if right_cell is None:
                            right_cell = CellExtended(worksheet.cell(row=i, column=j + 1))
                            virtual_worksheet[virtual_j + 1][virtual_i] = right_cell
                        cell.apply_right_neighbor(right_cell)
                        right_cell.apply_left_neighbor(cell)
                    else:
                        cell.apply_right_neighbor(None)
            parsed_cells = self.classify_cells(virtual_worksheet)
            column_metadata = ColumnMetaData(file_name=self.file_path, sheet_name=sheet_name)
            for new_col in ExtractHelper.extract_columns(parsed_cells, column_metadata):
                new_col.set_type()
                columns.append(new_col)
        return columns
